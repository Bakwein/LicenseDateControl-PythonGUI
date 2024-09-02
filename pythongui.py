import sys
import random
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QDateTimeEdit, QLabel, QPushButton, QLineEdit, QGridLayout, QListWidget, QTableWidget, QHeaderView, QTableWidgetItem, QFileDialog, QDateTimeEdit, QMessageBox, QSystemTrayIcon, QMenu
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt6.QtCore import QDateTime, Qt, QDate, QFile, QTextStream
from PyQt6.QtGui import QColor, QIcon, QAction
import sqlite3
from datetime import datetime
import pandas as pd
from plyer import notification
import os
import shutil
from qt_material import apply_stylesheet
'''
24.08

BİTTİ1 -1 AYDAN AZ SÜRE VARSA SARI 1 HAFTADAN AZ VARSA KIRMIZI 

BİTTİ2 - isim kontrolü olacak aynıysa önceki düzenlenecek farklıysa dict'e eklenecek ama bu zaten normal dicte ekleme ile de otomatik güncelleniyor olabilir buna bir bak

BİTTİ3 - EXCEL RENKLİ BACKGROUNDA YAZI RENGİ KONTROL

25.08

BİTTİ4- VERİYİ TUTULACAK HALE GETİRME (DÜZENLE SİL VERİ TABANI GÜNCELLENECEK)

BİTTİ5- 30 SINIRINA EKSTRA DURUMLAR EKLENECEK -> y_ olayından çözülebilir -> oldu ama kontrol lazim

BİTTİ6-SEARCH GETİRİLECEK

BİTTİ7- FİLTRELEME OLAYLARI GETİRİLEBİLİR - İSİME GÖRE, BİTİŞ TARİHİNE GÖRE BUTON

BİTTİ8- SİLDİKTEN, DÜZENLEDİKTEN SONRA, ARA YAPTIKTAN BİLDİRİM GİTMESİN

BİTTİ9-FİLTERLARDAN SONRA EXCEL TABLOSU NASIL ELDE EDİLİYOR KONTROL LAZİM

26.08

BİTTİ10- DÜZENLE KISMINDA HATA KISMI DÜZENLENECEK

BİTTİ11-DÜZENLE YAPILANCA VERİLER ARA KISMINA GELECEK

BİTTİ 12 - FİLTRELEDİKTEN SONRA TABLO BOYUTU KISMINA BAK?

BİTTİ 13 - SIRALA BUTONLARI AKTİFLİK GÖSTERİLEBİLİR

BİTTİ 14 GİBİ- EXCEL DOSYASI YÜKLE GETİRİLECEK

BİTTİ 15 GİBİ - DB DOSYASI YÜKLE

tablo uzunluk ayari ayarlandi

27.08

BİTTİ 16 - SİLME VE DÜZENLEME DURUMLARI İÇİN YEDEK DB ARADA ALİNABİLİR. -> butonlarla silme de eklendi

BİTTİ 17- TABLO HİZALAMA BİTTİ

BİTTİ 18- GİRİŞ İÇİN DB OLUŞTURULUYOR

BİTTİ 19- hata kontrolleri eklendi exception kontrolleri için try-catch kullanildi

28.08

BİTTİ 20 - tema geçişleri eklendi

BİTTİ 21 - silinmeden önce messagebox eklendi

BİTTİ 22- DÜZENLE VE SİLİNCE ELEMANLAR KALDIĞI YERDEN DEVAM EDİYOR

--------------

garip bir bug var bazen edit yapıldığına sol üstteki hücre gidiyor bu da hataya bus error'e sebep oluyor

ALTA DA BAK!

videolar tamam sadece bildirim kısmı eklenecek onu da windowsda yaparım ayrıca bir pdf hazırlayabilirm kullanım için -> haftasonu işi

'''

'''
EXE HALE GETİRME 
pip install pyinstaller
pyinstaller --onefile --noconsole pythongui.py

pyinstaller --onefile --hidden-import=PyQt6.QtCore --hidden-import=PyQt6.QtGui --hidden-import=PyQt6.QtWidgets pythongui.py

pyinstaller --onefile --hidden-import=PyQt6.QtCore --hidden-import=PyQt6.QtGui --hidden-import=PyQt6.QtWidgets --hidden-import plyer.platforms.win.notification  --noconsole pythongui.py

pyinstaller --onefile --hidden-import=PyQt6.QtCore --hidden-import=PyQt6.QtGui --hidden-import=PyQt6.QtWidgets --hidden-import=plyer.platforms.win.notification --hidden-import=plyer.platforms.macosx.notification --noconsole --icon=ico.ico pythongui.py

'''
'''
1. Başlatma Klasörüne Kısayol Ekleyin:
.exe dosyanıza sağ tıklayın ve Kısayol Oluştur seçeneğini seçin.
Windows + R tuşlarına basarak Çalıştır penceresini açın.
Çalıştır penceresine shell:startup yazın ve Enter tuşuna basın. Bu, Başlangıç (Startup) klasörünü açacaktır.
Oluşturduğunuz .exe dosyasının kısayolunu bu klasöre sürükleyip bırakın.


2.Kısayol Özelliklerini Düzenleme:

Oluşturduğunuz kısayola sağ tıklayın ve "Özellikler" seçeneğini seçin.
Açılan pencerede "Kısayol" sekmesine gidin.
"Çalıştır" seçeneğinin yanında bulunan açılır menüden "Simge Durumuna Küçültülmüş" seçeneğini seçin.
"Uygula" ve ardından "Tamam" butonlarına tıklayarak değişiklikleri kaydedin.

https://www.tenforums.com/tutorials/57690-create-elevated-shortcut-without-uac-prompt-windows-10-a.html !!!!!!

3.resim eklemeyi yap!!!

'''

'''
import pypyodbc as odbc #pip install pypyodbc

con = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};' #
                         'SERVER=server_adı;'
                         'DATABASE=veritabanı_adı;'
                         'UID=kullanıcı_adı;'
                         'PWD=parola')

#DRIVER={ODBC Driver 17 for SQL Server}; =>  Bu kısım, SQL Server veritabanına bağlanmak için kullanılacak ODBC (Open Database Connectivity) sürücüsünü belirtir.  SQL Server 2008 ve üstü sürümleri destekler.         

#SERVER=server_adı;

Açıklama: Bu kısım, SQL Server'ın bulunduğu sunucunun adını veya IP adresini belirtir.Eğer SQL Server yerel (localhost) olarak çalışıyorsa, buraya localhost veya . (nokta) yazabilirsiniz. Eğer SQL Server uzak bir sunucuda çalışıyorsa, buraya o sunucunun adı veya IP adresi yazılmalıdır. Örneğin, 192.168.1.100 gibi.

#DATABASE=veritabanı_adı; 
bağlanmak istediğiniz SQL Server veritabanının adını belirtir.
veritabanı_adı: Bağlanmak istediğiniz veritabanının adını buraya yazmalısınız. Örneğin, DATABASE=myDatabase; şeklinde.

con = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                     'SERVER=localhost;'
                     'DATABASE=MyDatabase;'
                     'UID=myUsername;'
                     'PWD=myPassword')

!DATABASE=MyDatabase; kısmı, SQL Server'da oluşturduğunuz veritabanının adını temsil eder. SQL Server'da veritabanları dosya bazlı değil, sunucu üzerinde tanımlanmış isimlerle yönetilir. Yani, MyDatabase SQL Server üzerinde oluşturduğunuz bir veritabanının adıdır, ve bu veritabanı birden fazla tabloyu barındırabilir. Oluşturmak için  SQL Server Management Studio (SSMS) gibi bir araç kullanarak yapabilirsiniz.

1.SQL Server'a Bağlanın
SQL Server Management Studio'yu (SSMS) başlatın.

2. Yeni Bir Veritabanı Oluşturun
SSMS'de "Object Explorer" penceresinde sunucuya sağ tıklayın ve "New Database" seçeneğini seçin.
Açılan pencerede veritabanınıza bir ad verin. Örneğin, tablo_verileri.

'''


'''
Sorunlar/Sorular /Soylenecekler:

SORUNLAR

SORULAR
-10 gün <= mi < mi?

'''

class DateTimePicker(QWidget):
    def __init__(self):
        super().__init__()
        self.resize(900, 500) #boyut
        self.showMinimized()
        self.dct = dict()
        self.colNo = 0
        self.rowNo = 0
        self.isDatePressed = True
        self.isEditModeOn = True
        self.editDict = {}

        '''
        self.setWindowFlags(Qt.WindowType.Tool | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        '''
        

        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(QIcon("ico.ico"))

        tray_menu = QMenu()

        #çıkış
        exit_action = QAction("Çıkış", self)
        exit_action.triggered.connect(self.close)
        tray_menu.addAction(exit_action)

        #menü sistem tepkisine ataniyor
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()

        #baslangic
        self.hide()

        #cift tik
        self.tray_icon.activated.connect(self.show_window)



        self.initUI()
        
    def initUI(self):
        #klasörler
        current_dir = os.getcwd()
        db_klasor = '/db_yedekleme'
        try:
            if not os.path.exists(current_dir + db_klasor):
                os.makedirs(current_dir + db_klasor)

            excel_klasor = '/excel_dosyalari'
            if not os.path.exists(current_dir + excel_klasor):
                os.makedirs(current_dir + excel_klasor)
        except Exception as e:
            self.hataLabel.setText("Klasörler oluşturulurken hata meydana geldi.")
        
        #database
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            #table
            #3 ' olacak
            c.execute('''
                  CREATE TABLE IF NOT EXISTS tablo_verileri (
                  isim TEXT,
                  start TEXT,
                  end TEXT
                  ) ''')
            c.execute('''
                    CREATE TABLE IF NOT EXISTS date_table (tarih TEXT
                      )
            ''')
            c.execute('''
                      CREATE TABLE IF NOT EXISTS
                      theme_table (themeType TEXT
                      )
                      ''')
            #tablo doldurma
            c.execute('SELECT * FROM tablo_verileri')
            veriler = c.fetchall()
            con.close()
        except Exception as e:
            #self.hataLabel.setText("Hata: Veri Tabanına Bağlanma Hatası"),
            return
        # #self.dct[isim_str] = {'name': isim_str, 'start': start_date_str, 'end': ending_date_str}
        for veri in veriler:
            self.getAnaDict()[veri[0]] = {'name': veri[0], 'start': str(veri[1]), 'end': str(veri[2]) }

        # giriş .db adında dosya oluşturma her girişte güncelleme
        try:
            '''
            con = sqlite3.connect('giris_tablo.db')
            c = con.cursor()
            c.execute(''
                  CREATE TABLE IF NOT EXISTS tablo_verileri (
                  isim TEXT,
                  start TEXT,
                  end TEXT
                  ) '')
            c.execute("DELETE FROM tablo_verileri")
            for key,value in self.getAnaDict().items():
                c.execute(''
                INSERT INTO tablo_verileri (isim, start, end)
                VALUES (?,?,?)
                '',(key,value['start'], value['end']))
            con.commit()
            con.close()
            '''

        except Exception as e:
            self.hataLabel.setText("Hata: Giriş Veri Tabanına Bağlanma Hatası")
            return

        #excel ve db dosyaları 1 ay geçmişse temizle

        #colNum = 5
        self.colNo = 5
        #rowNum = 34 # bu  sonra len olarak belirlenecek
        self.rowNo = len(veriler)

        #layout = QVBoxLayout()
        layout = QGridLayout()

        # Labeller
        self.isim = QLabel("İsim", self)
        self.isim.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")
        self.label = QLabel("Başlangıç Tarihi:", self)
        self.label.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")
        self.label2 = QLabel("Bitiş Tarihi:", self)
        self.label2.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")
        #self.listLabel = QLabel("Listeler:", self)

        #label2 yerini değiştirme

        self.uyari = QLabel("""*Database veya Excel Verisi Yüklediğinizde Eski Database "db_yedekleme" Klasöründe Yedeklenir""", self)
        self.uyari.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold; font-style: italic;")

        #isim giris
        self.isimTextbox = QLineEdit("")
        self.isimTextbox.setPlaceholderText("Lütfen İsmi Giriniz")

        # begging_date oluşturma
        self.beggining_date = QDateTimeEdit(self)
        self.beggining_date.setCalendarPopup(True) # açılır pencere
        self.beggining_date.setDateTime(QDateTime.currentDateTime())
        self.beggining_date.setDisplayFormat("dd.MM.yyyy")
        self.beggining_date.setFixedSize(110,40)

        #ending_date
        self.ending_date = QDateTimeEdit(self)
        self.ending_date.setCalendarPopup(True) # açılır pencere
        self.ending_date.setDateTime(QDateTime.currentDateTime())
        self.ending_date.setDisplayFormat("dd.MM.yyyy")
        self.ending_date.setFixedSize(110,40)
        #buton

        self.button = QPushButton("Ekle", self)
        #sinyal olayi 
        #self.button.setStyleSheet("background-image: linear-gradient(144deg,#AF40FF, #5B42F3 50%,#00DDEB);")
        self.button.clicked.connect(self.butonBasildi)

        self.listeGoster = QPushButton("*Excel Dosyası Yükle", self)
        self.listeGoster.clicked.connect(self.excelImport)

        self.dbEkle = QPushButton("*Yeni Database Yükle", self)
        self.dbEkle.clicked.connect(self.newDatabase)

        self.listeDuzenle = QPushButton("Tabloyu Excel Olarak Çıktı Al", self)
        self.listeDuzenle.clicked.connect(self.excelTabloOutput)

        self.listeAll = QPushButton("Tüm Verileri Excel Olarak Çıktı Al", self)
        self.listeAll.clicked.connect(self.excelOutput)
        
        #tablo
        self.table = QTableWidget()
        self.table.setRowCount(self.rowNo)
        self.table.setColumnCount(self.colNo)
        # isim - bas - bit - sil - düzenle (son 2 buton)
        self.table.setHorizontalHeaderLabels(['İsim', 'Başlangıç Tarihi', 'Bitiş Tarihi', 'Düzenle', 'Sil'])

        #header ayari
        self.headerAyar()
        
        #self.button.clicked.connect(self.butonBasildi)
        self.silLabel = QLabel("Dosyaları Temizleme",self)
        self.silLabel.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")

        self.yedeklemeDosyaSil = QPushButton(".db Dosyalarını Temizle", self)
        self.yedeklemeDosyaSil.clicked.connect(self.yedeklemeDosyaSilFunc1)

        self.yedeklemeDosyaSil2 = QPushButton(".xlsx Dosyalarını Temizle", self)
        self.yedeklemeDosyaSil2.clicked.connect(self.yedeklemeDosyaSilFunc2)

        self.yedeklemeDosyaSil3 = QPushButton("Tüm Dosyaları Temizle", self)
        self.yedeklemeDosyaSil3.clicked.connect(self.yedeklemeDosyaSilFunc3)
        
        #Search elemenleri
        self.searchLabel = QLabel("İsim Ara", self)
        self.searchLabel.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")
        self.searchTextbox = QLineEdit("")
        self.searchTextbox.setPlaceholderText("Lütfen Arayacağınız İsmi Giriniz")
        self.searchButton = QPushButton("Tabloda Ara", self)
        self.searchButton.clicked.connect(self.aramaBasildi)
        self.searchButton2 = QPushButton("Tümünde Ara", self)
        self.searchButton2.clicked.connect(self.aramaBasildi2)
        
        self.resetButton1 = QPushButton("Aramayı Temizle", self)
        self.resetButton1.clicked.connect(self.aramaTemizlemeBasildi)

        #Filtreleme 
        self.filterLabel = QLabel("Filtreleme/Sıralama", self)
        self.filterLabel.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")
        self.dateSort = QPushButton("Kalan Güne Göre Sırala", self)
        self.dateSort.clicked.connect(self.dateSortFunc)
        self.nameSort = QPushButton("İsme Göre Sırala", self)
        self.nameSort.clicked.connect(self.NameSortFunc)

        #bunlara da date sort lazim
        self.onlylast10 = QPushButton("Son 10 Gün", self) 
        self.onlylast10.clicked.connect(self.son10Goruntule)
        self.between10to30 = QPushButton("Son 1 Ay", self)
        self.between10to30.clicked.connect(self.between1030Basildi)
        self.normal = QPushButton("1 Aydan Fazla", self)
        self.normal.clicked.connect(self.normalBasildi)
        self.all = QPushButton("Tümünü Görüntüle", self)
        self.all.clicked.connect(self.aramaTemizlemeBasildi)

        #Hata yazma label'ı
        self.hataLabel = QLabel("", self)
        self.hataLabel.setStyleSheet("color: darkred; font-size:15px; font-weight: bold; font-style: italic;")

        #düzenleme modu
        self.editMode = QPushButton("Düzenleme/İnceleme Modunu Aç/Kapa", self)
        self.editMode.clicked.connect(self.editModeBasildi)

        #uyari
        self.uyari1 = QLabel("UYARI: Dosya fazlalığı problemi yaşamamanız için sıklıkla dosyaları temizlemeniz önerilir!", self)
        self.uyari1.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold; font-style: italic;")

        #son_tarih
        self.sonTarih = QLabel("Son Silinme Tarihi: Daha Önce Silinmemiş!", self)
        self.sonTarih.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold; font-style: italic;")
        self.tarihi_al()
        #tablo ilk doldurma
        self.tabloElemanlariEkle(True)
        self.dateSortFunc()
        self.editModeBasildi()

        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        #theme sayisini al
        self.getThemeNo()
        #theme sayisini isle
        self.applyTheme()

        #theme butonu işle
        self.themeButton = QPushButton("Tema Değiştir "+ str(self.themeNum+1) + "/3" ,self)
        self.themeButton.clicked.connect(self.themeButtonBasildi)

        self.labelTemaBilgi = QLabel("""Tema 1: Default Renkli Tablo\nTema 2: Koyu Renksiz Tablo\nTema 3: Açık Renksiz Tablo""", self)
        self.labelTemaBilgi.setStyleSheet("color: #1565C0; font-size:15px; font-weight: bold;")

        # Layout'a widgetları ekleme
        layout.addWidget(self.isim, 0, 0)
        layout.addWidget(self.isimTextbox, 1 , 0,1, 2)
        layout.addWidget(self.label, 2, 0)
        layout.addWidget(self.beggining_date, 3, 0)
        layout.addWidget(self.label2, 2, 1)
        layout.addWidget(self.ending_date, 3, 1)
        layout.addWidget(self.button, 4, 0, 1,1)
        #araya belki bir çizgi??

        #layout.addWidget(self.listLabel, 0,3 )
        layout.addWidget(self.dbEkle, 0,3)
        layout.addWidget(self.listeGoster, 0,4)
        layout.addWidget(self.listeDuzenle, 0, 5)
        layout.addWidget(self.listeAll, 0 ,6)
        #layout.addWidget(self.liste, 2, 3, 4, 2)
        layout.addWidget(self.table, 2, 3, 5, 5 )
        layout.addWidget(self.uyari, 1,3, 1,3)
        #arama yerlestirme
        layout.addWidget(self.searchLabel, 5, 0)
        layout.addWidget(self.searchTextbox, 6, 0,1,2)
        layout.addWidget(self.searchButton, 7,0)
        layout.addWidget(self.searchButton2, 8,0)
        layout.addWidget(self.resetButton1, 7,1)

        layout.addWidget(self.filterLabel, 7, 3)
        layout.addWidget(self.dateSort, 8, 3)
        layout.addWidget(self.nameSort, 8, 4)
        layout.addWidget(self.onlylast10, 9, 3)
        layout.addWidget(self.between10to30, 9, 4)
        layout.addWidget(self.normal, 9, 5)
        layout.addWidget(self.all, 9,6)
        layout.addWidget(self.hataLabel,7,5,1,2)
        layout.addWidget(self.editMode,1,6)

        # Ana pencereye layout'u ekleme
        

        #Silme Layoutları
        layout.addWidget(self.silLabel, 10,3)
        layout.addWidget(self.yedeklemeDosyaSil, 11, 3)
        layout.addWidget(self.yedeklemeDosyaSil2, 11, 4)
        layout.addWidget(self.yedeklemeDosyaSil3, 11, 5)
        layout.addWidget(self.uyari1, 10,4, 1, 3)
        layout.addWidget(self.sonTarih, 11,6)

        layout.addWidget(self.themeButton, 10,0)
        layout.addWidget(self.labelTemaBilgi, 9,0,1,2)

        self.setLayout(layout)

        # Pencereyi ayarlama
        self.setWindowTitle('Toyotetsu-Tarih Kontrol GUI')
        self.setWindowIcon(QIcon('ico.ico'))
        self.show()
        
    def butonBasildi(self):
        noti = False
        temp_isDatePressed = self.isDatePressed

        #texti alma
        isim = self.isimTextbox
        if not isim:
            self.hataLabel.setText("Hata: İsime erişilemiyor.")
            return
        isim_str = isim.text()
        if isim_str == "":
            self.hataLabel.setText("Hata: İsim boş olamaz!")
            return

        table_size = self.table.rowCount()

        #boyle bir isim var mi
        xd = self.tabloVerileriEldeEt()
        xd_keys = xd.keys()
        if isim_str not in xd_keys:
            if (table_size + 1 > self.rowNo):
                self.rowNo += 1
                self.table.setRowCount(self.rowNo)

        #başlangıç tarihini alma
        begging_date = self.beggining_date.dateTime()

        day_zero1 = f'{begging_date.date().day():02}'
        month_zero1 = f'{begging_date.date().month():02}'
        start_date_str = day_zero1 + '.' + month_zero1 + '.' + str(begging_date.date().year())

        #bitiş tarihini alma
        ending_date = self.ending_date.dateTime()

        day_zero = f'{ending_date.date().day():02}'
        month_zero = f'{ending_date.date().month():02}'
        ending_date_str = day_zero + '.'+month_zero + '.' + str(ending_date.date().year())

        dif_ = self.calculateDif(ending_date_str)
        if(dif_.days <= 10):
            noti = True
        
        ret = self.kontrolluEkle(isim_str=isim_str, start_date_str=start_date_str, ending_date_str=ending_date_str, noti=noti)

        self.isDatePressed = temp_isDatePressed
        #self.dct[isim_str] = {'name': isim_str, 'start': start_date_str, 'end': ending_date_str}
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()

        if not ret:
            self.hataLabel.setText("Hata: Lütfen verileri doğru formatta giriniz!")
            return 
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute('DELETE FROM tablo_verileri')
            for key,value in self.getAnaDict().items():
                #3 ' unutma
                c.execute('''
                INSERT INTO tablo_verileri (isim, start, end)
                VALUES (?,?,?)
                ''',(key,value['start'], value['end']))
            con.commit()
            con.close()
        except:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 1!")
            return
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def excelImport(self):
        opt = QFileDialog.Option.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(self, "Excel Dosyasını Seç", "","Excel Files (*.xlsx);;All Files (*)", options=opt)
        if file_path:
            try:
                df = pd.read_excel(file_path, engine="openpyxl")
            except:
                self.hataLabel.setText("Hata: Excel Dosyasını Okurken Hata Oluştu.")
                return
            if len(df.columns) != 3:
                self.hataLabel.setText("Hata: Excel Tablo Formatı Yanlış") 
                return
            df.columns = ['name', 'start', 'end']
            all_elem = df.to_dict('records')
            new_dct = {}
            for elem in all_elem: 
                if not self.isValidDate(elem['start']) or not self.isValidDate(elem['end']):
                    self.hataLabel.setText("Hata: Excel Dosyasında Yanlış Tarih Formatı Bulundu.")
                    return
                new_dct[str(elem['name'])] = {'name': elem['name'], 'start': elem['start'], 'end': elem['end']}
            self.dct = new_dct
            self.setAnaDict(new_dct)
            self.tabloBoyutInput(len(self.getAnaDict()))
            #self.tabloTemizleme()
            self.tablo_guncelle(new_dct=self.getAnaDict(), noti=True)
            
            #self.editModeBasildi()
            now = datetime.now()
            date_time_str = now.strftime('%Y%m%d_%H%M%S')
            full_path = os.path.join(os.getcwd()+ f'/db_yedekleme/EXCEL_BACKUP_tablo_verileri_backup_{date_time_str}.db')
            shutil.copy('./tablo_verileri.db',full_path)

            try:
                con = sqlite3.connect('tablo_verileri.db')
                c = con.cursor()
                #şimdiki tablo_verileri'ni sakla!!!!!
           
                #name = f'tablo_verileri_{date_time_str}'
                c.execute('DELETE FROM tablo_verileri')
                for key,value in self.getAnaDict().items():
                #3 ' unutma
                    c.execute('''
                    INSERT INTO tablo_verileri (isim, start, end)
                    VALUES (?,?,?)
                    ''',(key,value['start'], value['end']))
                con.commit()
                con.close()
            except:
                self.hataLabel.setText("Hata: Veri Tabanı Hatası 2")
                return
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")
        self.dateSort.setStyleSheet("")
        self.nameSort.setStyleSheet("")

    def getTableLength(self):
        l = 0
        for row in range(self.table.rowCount()):
            item = self.table.item(row,0)
            if item is not None and item.text().strip():
                l+=1
        return l

    def tabloVerileriEldeEt(self):
        ret_dct = {}
        #print(self.table.rowCount(),"*",self.getTableLength())
        for row in range(self.getTableLength()):
            name_item = self.table.item(row, 0)
            start_date_item = self.table.item(row, 1)
            end_date_item = self.table.item(row, 2)
            #print(name_item, start_date_item, end_date_item, "*")
            if name_item == None or start_date_item == None or end_date_item == None:
                self.hataLabel.setText("Hata: Tablo elemanlarına erişilemiyor.")
                return
            name = name_item.text()
            start_date = start_date_item.text()
            end_date = end_date_item.text()
            ret_dct[name] = {'name': name, 'start': start_date, 'end': end_date}
        return ret_dct
    
    def excelOutput(self):
        ret_dict = self.getAnaDict() 
        if not self.isDatePressed:
            ret_dict = dict(sorted(ret_dict.items()))
        else:
            ret_dict = dict(sorted(ret_dict.items(), key=lambda item: self.calculateDif(item[1]['end'])))
        df = pd.DataFrame(ret_dict).T
        df.columns = ("İsim", "Başlangıç Tarihi", "Bitiş Tarihi")
        if df.size == 0:
            self.hataLabel.setText("Hata: Yazılacak Bilgi Yok!")
            return
        df.fillna('', inplace=True)
        filename, _ = QFileDialog.getSaveFileName(self, "Excel dosyasını kaydet", "", "Excel Files (*.xlsx);;All Files (*)")
        if filename:
            try:
                df.to_excel(filename, index=False)
                workbook = load_workbook(filename)
                sheet = workbook.active
                keys = list(ret_dict.keys())

                for y in range(len(ret_dict)):
                    #item_text = self.table.item(y,2).text()
                    item_text = self.getAnaDict()[keys[y]]['end']
                    #print(item_text, " boom!")
                    dif = self.calculateDif(item_text)
                    if dif.days <= 10:
                    #print("red")
                        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        for i in ['A', 'B', 'C']:
                            sheet[i + str(y+2)].fill = red_fill
                    elif dif.days < 30:
                        #print("yellow")
                        yellow_fill = PatternFill(start_color="FFFF33", end_color="FFFF33",fill_type="solid")
                        for i in ['A', 'B', 'C']:
                            sheet[i+str(y+2)].fill = yellow_fill
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Kolon harfini al
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # İhtiyaca göre genişliği ayarla
                    sheet.column_dimensions[column].width = adjusted_width
                workbook.save(filename)
            except Exception as ex:
                self.hataLabel.setText("Hata: Excel Çıktı Alınırken Hata Oluştu")
                return
        self.duzenle_kapali()
        self.hataLabel.setText("")

    def excelTabloOutput(self):
        tablo_dct = self.tabloVerileriEldeEt()
        if len(tablo_dct) == 0:
            self.hataLabel.setText("Hata: Tabloda veri yok!")
            return
        df = pd.DataFrame(tablo_dct).T
        df.columns = ("İsim", "Başlangıç Tarihi", "Bitiş Tarihi")
        if df.size == 0:
            return
        df.fillna('', inplace=True)
        filename, _ = QFileDialog.getSaveFileName(self, "Excel dosyasını kaydet", "", "Excel Files (*.xlsx);;All Files (*)")
        if filename:
            try:
                df.to_excel(filename, index=False)
                workbook = load_workbook(filename)
                sheet = workbook.active
                for y in range(len(tablo_dct)):
                    item_text = self.table.item(y,2).text()
                    dif = self.calculateDif(item_text)
                    if dif.days <= 10:
                        #print("red")
                        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        for i in ['A', 'B', 'C']:
                            #print(i + str(y), "boom")
                            sheet[i + str(y+2)].fill = red_fill
                    elif dif.days < 30:
                        #print("yellow")
                        yellow_fill = PatternFill(start_color="FFFF33", end_color="FFFF33",fill_type="solid")
                        for i in ['A', 'B', 'C']:
                            sheet[i+str(y+2)].fill = yellow_fill

                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Kolon harfini al
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # İhtiyaca göre genişliği ayarla
                    sheet.column_dimensions[column].width = adjusted_width
                workbook.save(filename)
            except Exception as ex:
                self.hataLabel.setText("Hata: Excel Çıktı Alınırken Bir Hata Oluştu.")
                return
        self.duzenle_kapali()
        self.hataLabel.setText("")

    def kontrolluEkle(self, isim_str, start_date_str, ending_date_str, noti):
        ret = True
        #kontroller olacak burada - sorun varsa false olsun
        if isim_str == "" or not self.isValidDate(start_date_str) or not self.isValidDate(ending_date_str):
            self.hataLabel.setText("Hata: Lütfen Doğru Formatta Tekrar Deneyiniz.")
            return False
        
        self.getAnaDict()[isim_str] = {'name': isim_str, 'start': start_date_str, 'end': ending_date_str}
        ret2 = self.tabloElemanlariEkle(noti)
        if not ret2:
            ret = False
        return ret

    def tabloElemanlariEkle(self, noti):
        #print(self.dct, type(self.dct))
        try:
            self.aramaTemizlemeBasildi()
            if len(self.getAnaDict()) != 0:
                y_ = 0
                #self.NameSortFunc()
                #print(self.getAnaDict(), "önce")
                if self.isDatePressed:
                    self.dateSortFunc() 
                else:
                    self.NameSortFunc()
                #print(self.getAnaDict(), "after")
                for key,value in self.getAnaDict().items():
                    self.table.setItem(y_, 0, QTableWidgetItem(key))
                    self.table.setItem(y_, 1, QTableWidgetItem(value["start"]))
                    self.table.setItem(y_, 2, QTableWidgetItem(value["end"]))

                    #düzenle butonu
                    but1 = QPushButton("Düzenle", self)
                    but1.clicked.connect(lambda _, r=3, c=y_: self.tabloButtonDuzenle(r,c))
                    self.table.setCellWidget(y_, 3, but1)

                    #silme butonu
                    but2 = QPushButton("Sil", self)
                    but2.clicked.connect(lambda _, r=4, c=y_: self.tabloButtonSil(r,c))
                    self.table.setCellWidget(y_, 4, but2)
                    y_+=1
                    if (y_ > self.rowNo):
                        self.rowNo = y_
                        self.table.setColumnCount(self.rowNo)
                self.table.resizeColumnsToContents()
                self.table.resizeRowsToContents()
                ret = self.tabloGunKontrol(self.getAnaDict(), noti)
                if not ret :
                    return False
                return True
        except Exception as ex:
            self.hataLabel.setText("Hata: Tabloya eleman eklenirken hata oluştu.")

    def toDate(str):
        return datetime.strptime(str, '%d.%m.%Y')
    
    def changeColorSort(self, isDate):
        if not self.dateSort:
            self.hataLabel.setText("Hata: Tarihe göre sıralama butonunda hata!")
            return
        if not self.nameSort:
            self.hataLabel.setText("Hata: İsme göre sıralam butonunda hata.")
            return
        
        if isDate:                
            self.dateSort.setStyleSheet("background-color: lightgreen; color:black; border-style: outset; border-width:2px; border-radius:5px; border-color: beige;")
            self.nameSort.setStyleSheet("")
            self.isDatePressed = True
        else:
            self.dateSort.setStyleSheet("")
            self.nameSort.setStyleSheet("background-color: lightgreen; color:black; border-style: outset; border-width:2px; border-radius:5px; border-color: beige; ")
            self.isDatePressed = False

    def dateSortFunc(self):
        self.changeColorSort(True)
        new_dct = self.tabloVerileriEldeEt()
        new_dct = dict(sorted(new_dct.items(), key=lambda item: self.calculateDif(item[1]['end'])))
        self.tabloTemizleme()
        self.tablo_guncelle(new_dct)
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def dateSortRet(self, dct):
        return dict(sorted(dct.items(), key=lambda item: self.calculateDif(item[1]['end'])))

    def NameSortFunc(self):
       self.changeColorSort(False)
       new_dct = self.tabloVerileriEldeEt()
       new_dct= dict(sorted(new_dct.items()))
       self.tabloTemizleme()
       self.tablo_guncelle(new_dct)
       self.duzenle_kapali()
       self.headerAyar()
       self.hataLabel.setText("")

    def nameSortRet(self,dct):
        return dict(sorted(dct.items()))

    def isValidDate(self, date):
        date_format = "%d.%m.%Y"
        try:
            datetime.strptime(date, date_format)
            return True
        except ValueError:
            return False

    def tabloButtonDuzenle(self, r, c):
        keys_list = list(self.editDict.keys())
        old_key = keys_list[r]
        name = self.table.item(r,0).text()
        bas_tarih = self.table.item(r,1).text()
        son_tarih = self.table.item(r,2).text()

        if not (self.isValidDate(bas_tarih) and self.isValidDate(son_tarih) and len(name) != 0):
            if len(name) == 0:
                self.hataLabel.setText("Hata: İsim boş olamaz!")
            elif not self.isValidDate(bas_tarih):
                self.hataLabel.setText("Hata: Başlangıç Tarihi Formatı Yanlış!")
            else:
                self.hataLabel.setText("Hata: Bitiş Tarihi Formatı Yanlış!")
            return 
        new_key = self.table.item(r,0).text()
        try:
            del self.getAnaDict()[old_key]
        except Exception as ex:
            self.hataLabel.setText("Hata: Silme İşleminde Hata Oluştu!")
            return
        self.getAnaDict()[new_key] = {'name': new_key, 
                             'start': bas_tarih,
                             'end': son_tarih
                             }
        '''
        temp_dict[new_key] = {'name': new_key, 
                             'start': bas_tarih,
                             'end': son_tarih
                             }
            '''
        r_ = self.tabloAyniGuncelle()
        self.tablo_guncelle(new_dct=r_)
        
        #Ara kısmını güncelle
        self.isimTextbox.setText(new_key)
        
        datime_obj1 = QDateTime.fromString(bas_tarih,"dd.MM.yyyy")
        self.beggining_date.setDateTime(datime_obj1)

        datetime_obj2 = QDateTime.fromString(son_tarih,"dd.MM.yyyy")
        self.ending_date.setDateTime(datetime_obj2)

        self.editDict = r_
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute('''UPDATE tablo_verileri
                       SET isim = ?, start = ?, end = ?
                       WHERE isim = ?
            ''',(new_key,bas_tarih,son_tarih,old_key ))
            con.commit()
            con.close()
        except Exception as ex:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 3!")
            return
        self.duzenle_acik()
        self.headerAyar()
        self.table.setCurrentCell(-1, -1)
       
    def tabloButtonSil(self, r, c):
        #dialog
        dlg = QMessageBox(self)
        dlg.setWindowTitle("Uyarı!")
        dlg.setText(f"Tablodaki {r+1}. eleman silinecek. Emin misiniz?")
        dlg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        dlg.setIcon(QMessageBox.Icon.Question)
        button = dlg.exec()
        if button == QMessageBox.StandardButton.No:
            return
    
        get_table = self.tabloVerileriEldeEt()
        name_ = self.table.item(r,0).text()
        try: 
            del self.getAnaDict()[name_]
            del get_table[name_]
        except Exception as ex:
            self.hataLabel.setText("Hata: Silme İşleminde Hata Oluştu 2!")
            return
        
        sorted_tablo_veri = None
        if self.isDatePressed:
            sorted_tablo_veri = self.dateSortRet(get_table)
        else:
            sorted_tablo_veri = self.nameSortRet(self.getAnaDict())
        self.tablo_guncelle(new_dct=sorted_tablo_veri)
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute(''' DELETE FROM tablo_verileri
                        WHERE isim = ?
                    ''',(name_,))
            con.commit()
            con.close()
        except Exception as ex:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 4!")
            return
        self.headerAyar()
        self.duzenle_kapali()
        self.table.setCurrentCell(-1, -1)

    def tabloTemizleme(self):
        try:
            # Aktif düzenleyiciyi kapat
            self.table.setCurrentCell(-1, -1)
            if self.table.isActiveWindow():
                self.table.closePersistentEditor(self.table.currentItem())
            self.table.clearContents()
            #self.table.clearContents()
            self.table.setCurrentCell(-1, -1)
        except Exception as ex:
            self.hataLabel.setText("Hata: Tablo Temizlenirken Hata Oluştu.")

    def calculateDif(self, get_item_date):
        date_format = "%d.%m.%Y"
        ret_date = datetime.strptime(get_item_date, date_format)
        date = ret_date.date()
        current_date = datetime.now().date()
        dif = date - current_date
        return dif

    def tabloGunKontrol(self, dct, noti):
        try:
            len_ =  len(dct)
            isNotificationRequired = False
            for e in range(len_):
                #print(e)
                get_item = self.table.item(e, 2)
                get_item_date = get_item.text()
                if not self.isValidDate(get_item_date):
                    return False
                dif = self.calculateDif(get_item_date=get_item_date)
                if dif.days <= 10:
                    # 10 günden az - hem boya hem uyarı olayını ayarla
                    isNotificationRequired = True
                    self.tabloKirmiziBoya(num=e)
                elif dif.days < 30:
                    self.tabloSariBoya(num=e)
        except Exception as ex:
            self.hataLabel.setText("Hata: Tablo Gün Kontrolünde Hata!")
            return False
        try:
            if isNotificationRequired and noti:
                notification.notify(
                title = "Azalmış Gün Uyarısı!",
                message = "10 günden az bir tarih bulundu!",
                app_name = "Tarih Kontrol Uygulaması",
                timeout = 5
            )
        except Exception as ex:
            self.hataLabel.setText("Bildirim Gönderirken Hata Oluştu.")
            return False
        return True

    def tabloSariBoya(self,num):
        for s in range(3):
            item = self.table.item(num, s)
            item.setBackground(QColor(255, 255, 51))
            item.setForeground(QColor(0,0,0))

    def tabloKirmiziBoya(self,num):
        for s in range(3):
            item = self.table.item(num,s)
            item.setBackground(QColor(255, 0, 51))
            item.setForeground(QColor(0,0,0))

    def tablo_yazdir(self):
        dct = self.tabloVerileriEldeEt()
        for key,value in dct.items():
            print(key,":",value)

    def tablo_guncelle(self, new_dct,noti=False):
        try:
            # Aktif düzenleyiciyi kapat
            self.table.setCurrentCell(-1, -1)
            self.table.closePersistentEditor(self.table.currentItem())
            #if len(new_dct) == 0:
            #    return
            y_ = 0
        
            #self.tabloTemizleme()
            self.rowNo = len(new_dct)
            self.table.setRowCount(self.rowNo)
            for key,value in new_dct.items():
                self.table.setItem(y_, 0, QTableWidgetItem(key))
                self.table.setItem(y_, 1, QTableWidgetItem(value["start"]))
                self.table.setItem(y_, 2, QTableWidgetItem(value["end"]))
                #düzenle
                but1 = QPushButton("Düzenle", self)
                but1.clicked.connect(lambda _, r=y_, c=3: self.tabloButtonDuzenle(r,c))
                self.table.setCellWidget(y_, 3, but1)
                #sil butonu
                but2 = QPushButton("Sil", self)
                but2.clicked.connect(lambda _, r=y_, c=4: self.tabloButtonSil(r,c))
                self.table.setCellWidget(y_, 4, but2)
                y_+=1

            self.table.resizeColumnsToContents()
            self.table.resizeRowsToContents()
            ret = self.tabloGunKontrol(new_dct, noti)
            self.headerAyar()
            return ret
        except Exception as ex:
            self.hataLabel.setText("Hata: Tablo Güncellenirken Hata Oluştu!")
            return
        
    def database_yazdir(self):
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute('SELECT * FROM tablo_verileri')
            rows = c.fetchall()
            for row in rows:
                print(row)
            con.close()
        except Exception as ex:
            self.hataLabel.setText("Hata: Tablo Yazdırılıken Hata Oluştu.")
            return

    def aramaBasildi(self):
        arama = self.searchTextbox
        if not arama:
            self.hataLabel.setText("Hata: Arama Kutusuna Erişilemiyor.")
        arama_text = arama.text()
        if arama_text == "":
            self.hataLabel.setText("Hata: Arama boş olamaz!")
            return
        tablo_dict = self.tabloVerileriEldeEt()
        if not len(tablo_dict):
            self.hataLabel.setText("Hata: Tablo boşken arama yapılamaz!")
            return
        self.tabloTemizleme()
        new_dict = {}
        for key, value in tablo_dict.items():
            if arama_text in key:
                new_dict[key]=value 
        self.tabloBoyutInput(len(new_dict))
        self.tablo_guncelle(new_dct=new_dict)
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def aramaBasildi2(self):
        arama_text = self.searchTextbox.text()
        if arama_text == "":
            self.hataLabel.setText("Hata: Arama boş olamaz!")
            return
        tablo_dict = self.getAnaDict()
        if not len(tablo_dict):
            self.hataLabel.setText("Hata: Veri olmadan arama yapılamaz!")
            return
        self.tabloTemizleme()
        new_dict = {}
        for key, value in tablo_dict.items():
            if arama_text in key:
                new_dict[key]=value
        self.tabloBoyutInput(len(new_dict))
        self.tablo_guncelle(new_dct=new_dict)
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def aramaTemizlemeBasildi(self):
        if self.isDatePressed:
            #self.dateSortFunc()
            self.setAnaDict(self.dateSortRet(self.getAnaDict()))
        else:
            #self.NameSortFunc()
            self.setAnaDict(self.nameSortRet(self.getAnaDict()))
        self.tabloBoyutReset()
        self.tabloTemizleme()
        self.tablo_guncelle(self.getAnaDict())
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def son10Goruntule(self):
        ret_dict = {}
        tablo_dict = self.getAnaDict()
        for key,value in tablo_dict.items():
            ending_date = value['end']
            dif = self.calculateDif(ending_date)
            #print(dif.days)
            if dif.days <= 10:
                #print("AAAAA")
                ret_dict[key]=value
        self.tabloBoyutInput(len(ret_dict))
        self.tablo_guncelle(ret_dict)
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")
        
    def between1030Basildi(self):
        ret_dict = {}
        tablo_dict = self.getAnaDict()
        for key,value in tablo_dict.items():
            ending_date = value['end']
            dif = self.calculateDif(ending_date)
            #print(dif.days)
            if dif.days > 10 and dif.days <= 30:
                #print("AAAAA")
                ret_dict[key]=value
        self.tabloBoyutInput(len(ret_dict))
        self.tablo_guncelle(ret_dict)
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def normalBasildi(self):
        ret_dict = {}
        tablo_dict = self.getAnaDict()
        for key,value in tablo_dict.items():
            ending_date = value['end']
            dif = self.calculateDif(ending_date)
            #print(dif.days)
            if dif.days > 30:
                ret_dict[key]=value
        self.tabloBoyutInput(len(ret_dict))
        self.tablo_guncelle(ret_dict)
        if self.isDatePressed:
            self.dateSortFunc()
        else:
            self.NameSortFunc()
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")

    def newDatabase(self):
        try:
            opt = QFileDialog.Option.ReadOnly
            file_path, _ = QFileDialog.getOpenFileName(self, "Veritabanı dosyasını seç", "", "Database Files (*.db);;All Files (*)", options=opt)
            if file_path:
                con = sqlite3.connect(file_path)
                df = pd.read_sql_query("SELECT * FROM tablo_verileri", con)
                con.close()
                self.tabloTemizleme()
                df.columns = ['name', 'start', 'end']
                all_elem = df.to_dict('records')
                new_dct = {}
                for elem in all_elem:
                    if not self.isValidDate(elem['start']) or not self.isValidDate(elem['end']):
                        self.hataLabel.setText("Hata: Excel Dosyasında Yanlış Tarih Formatı Bulundu.")
                        return
                    new_dct[elem['name']] = {'name': elem['name'], 'start': elem['start'], 'end': elem['end']}
                #self.dct = new_dct
                self.setAnaDict(new_dct)
                self.tabloBoyutInput(len(self.getAnaDict()))
                self.tablo_guncelle(new_dct=self.getAnaDict(), noti=True)

                now = datetime.now()
                date_time_str = now.strftime('%Y%m%d_%H%M%S')
                full_path = os.path.join(os.getcwd()+ f'/db_yedekleme/DB_BACKUP_tablo_verileri_backup_{date_time_str}.db')
                shutil.copy('./tablo_verileri.db',full_path)

                try:
                    con = sqlite3.connect("tablo_verileri.db")
                    c = con.cursor()

                    c.execute('DELETE FROM tablo_verileri')
                    for key,value in self.getAnaDict().items():
                        c.execute('''
                        INSERT INTO tablo_verileri (isim, start, end)
                        VALUES (?,?,?)
                        ''',(key,value['start'], value['end']))
                    con.commit()
                    con.close()
                except Exception as ex:
                    self.hataLabel.setText("Hata: Veri Tabanı Hatası 5")
                    return       
        except:
            self.hataLabel.setText("Hata: Yeni Database Oluşturulurken Hata Oluştu!")
            return
        self.duzenle_kapali()
        self.headerAyar()
        self.hataLabel.setText("")
        self.dateSort.setStyleSheet("")
        self.nameSort.setStyleSheet("")

    def tabloBoyutGuncelle(self):
        boyut = self.getTableLength()
        self.table.setRowCount(boyut)

    def tabloBoyutReset(self):
        boyut = len(self.getAnaDict())
        self.table.setRowCount(boyut)

    def tabloBoyutInput(self, boyut):
        self.table.setRowCount(boyut)
        for row in range(boyut):
            for col in range(self.table.columnCount()):
                #print(row, col)
                if not self.table.item(row,col):
                    self.table.setItem(row,col,QTableWidgetItem(""))

    def duzenle_acik(self):
        for row in range(self.table.rowCount()):
            but = self.table.cellWidget(row,3)
            if but:
                but.setEnabled(True)
            else:
                self.hataLabel.setText("Hata: Tablo Buton Erişim Hatası")
                return

    def duzenle_kapali(self):
        for row in range(self.table.rowCount()):
            but = self.table.cellWidget(row,3)
            if but:
                but.setEnabled(False)
            else:
                self.hataLabel.setText("Hata: Tablo Buton Erişim Hatası")
                return

    def button_mode(self, bool):
        if not self.button or not self.searchButton or not self.searchButton2 or not self.resetButton1 or not self.dbEkle or not self.listeGoster or not self.listeGoster or not self.listeDuzenle or not self.listeAll or not self.dateSort or not self.nameSort or not self.onlylast10 or not self.between1030Basildi or not self.normal or not self.all:
            self.hataLabel.setText("Hata: Buton Modu Değiştirilirken Hata")
            return
        self.button.setEnabled(bool)
        self.searchButton.setEnabled(bool)
        self.searchButton2.setEnabled(bool)
        self.resetButton1.setEnabled(bool)
        self.dbEkle.setEnabled(bool)
        self.listeGoster.setEnabled(bool)
        self.listeDuzenle.setEnabled(bool)
        self.listeAll.setEnabled(bool)
        self.dateSort.setEnabled(bool)
        self.nameSort.setEnabled(bool)
        self.onlylast10.setEnabled(bool)
        self.between10to30.setEnabled(bool)
        self.normal.setEnabled(bool)
        self.all.setEnabled(bool)
        self.yedeklemeDosyaSil.setEnabled(bool)
        self.yedeklemeDosyaSil2.setEnabled(bool)
        self.yedeklemeDosyaSil3.setEnabled(bool)
        for row in range(self.table.rowCount()):
            but = self.table.cellWidget(row,3)
            if but:
                but.setEnabled(not bool)
            else:
                self.hataLabel.setText("Hata: Tablo Buton Erişim Hatası")
                return

    def editModeBasildi(self):
        if self.isEditModeOn:
            self.isEditModeOn = False
            self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self.button_mode(True)
            self.headerAyar()
            self.editMode.setStyleSheet("")
        else:
            self.isEditModeOn = True
            self.editMode.setStyleSheet("background-color: lightgreen; color:black; border-style: outset; border-width:2px; border-radius:5px; border-color: beige;")
            self.table.setEditTriggers(QTableWidget.EditTrigger.AllEditTriggers)
            self.button_mode(False)
            self.editDict = self.tabloVerileriEldeEt()

    def yedeklemeDosyaSilFunc1(self):
        self.pathTemizleme("./db_yedekleme")
        self.hataLabel.setText("")
        self.tarih_guncelle()
        self.tarihi_al()

    def yedeklemeDosyaSilFunc2(self):
        self.pathTemizleme("./excel_dosyalari")
        self.hataLabel.setText("")
        self.tarih_guncelle()
        self.tarihi_al()

    def yedeklemeDosyaSilFunc3(self):
        self.yedeklemeDosyaSilFunc1()
        self.yedeklemeDosyaSilFunc2()
        self.hataLabel.setText("")

    def pathTemizleme(self, path):
        fn = ""
        if path == "./db_yedekleme":
            fn = ".db"
        else:
            fn = ".xlsx"

        if os.path.exists(path):
            for filename in os.listdir(path):
                file_path = os.path.join(path,filename)
                try:
                    if (os.path.isfile(file_path) or os.path.islink(file_path)) and filename.endswith(fn):
                        os.unlink(file_path)
                except Exception as e:
                    #print("Hata")
                    self.hataLabel.setText(f"{fn} dosyaları silinirken bir hata meydana geldi")
                    return

    def headerAyar(self):
        self.header = self.table.horizontalHeader()
        self.header.setStretchLastSection(True)
        for col in range(self.colNo):
            self.header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)

    def tarih_guncelle(self):
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute('''
                DELETE FROM date_table
                ''')
            now = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
            #print(now, "!*!*!*!")
            c.execute('INSERT INTO date_table (tarih) VALUES (?)', (now,))
            con.commit()
            con.close()
        except Exception as e:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 6" + " " + str(e))
            
    def tarihi_al(self):
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute("SELECT tarih FROM date_table")
            veriler = c.fetchall()
            if len(veriler) > 1:
                self.hataLabel.setText("Hata: Tarih Veri Tabanı Tablosunda 1'den Fazla Değer Olamaz.")
                c.execute("DELETE FROM date_table")
                con.commit()
            elif len(veriler) == 1:
                self.sonTarih.setText("Son Silinme Tarihi: " + str(veriler[0][0])) 
            con.close()
        except Exception as e:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 7")
        
    def getThemeNo(self):
        try:
            con = sqlite3.connect('tablo_verileri.db')
            c = con.cursor()
            c.execute("SELECT themeType from theme_table")
            veriler = c.fetchall()
            if len(veriler) > 1:
                self.hataLabel.setText("Hata: Tema Veri Tabanı Tablosunda 1'den Fazla Değer Olamaz.")
                c.execute("DELETE FROM theme_table")
                con.commit()
            elif len(veriler) == 1:
                self.themeNum = int(veriler[0][0])
            elif len(veriler) == 0:
                self.themeNum = 0
            con.close()
        except Exception as e:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 8")

    def applyTheme(self):
        if self.themeNum == 0:
            app.setStyleSheet("QWidget {color: black; }")
        elif self.themeNum == 1:
            apply_stylesheet(app, theme='dark_blue.xml')
        elif self.themeNum == 2:
            apply_stylesheet(app, theme='light_blue.xml')
    
    def setThemeNum(self,no):
        self.themeNum = no
    
    def themeButtonBasildi(self):
        self.themeNum += 1
        if self.themeNum == 3:
            self.themeNum = 0
        self.themeButton.setText("Tema Değiştir "+ str(self.themeNum+1) + "/3")
        #veritabanı güncelle
        try:
            con = sqlite3.connect("tablo_verileri.db")
            c = con.cursor()
            c.execute("DELETE FROM theme_table")
            c.execute("INSERT INTO theme_table (themeType) VALUES (?)", (self.themeNum,))
            con.commit()
            c.close()
        except Exception as e:
            self.hataLabel.setText("Hata: Veri Tabanı Hatası 9")
        self.applyTheme()
        self.hataLabel.setText("")
        self.duzenle_kapali()
        self.headerAyar()
    
    def getAnaDict(self):
        return self.dct

    def setAnaDict(self, dct_):
        self.dct = dct_

    def tabloAyniGuncelle(self):
        tablo_veri = self.tabloVerileriEldeEt()
        sorted_tablo_veri = None
        if self.isDatePressed:
            sorted_tablo_veri = self.dateSortRet(tablo_veri)
        else:
            sorted_tablo_veri = self.nameSortRet(tablo_veri)
        return sorted_tablo_veri
    
    def show_window(self, reason):
        if reason == QSystemTrayIcon.ActivationReason.Trigger:
            self.showNormal()
            #self.activateWindow()
    
if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        ex = DateTimePicker()
        sys.exit(app.exec())
    except Exception as ex:
        print(ex, "exception detected.")
